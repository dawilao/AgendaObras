"""
Serviços de importação e exportação de cotações (PLO, template, ABC).
Todas as funções de exportação retornam bytes prontos para ui.download().
"""

import io
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from db.cotacoes_repo import CotacoesDatabase
from core.error_logger import log_error


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalizar_cabecalho(valor: Optional[str]) -> str:
    return (valor or '').strip().upper().replace('\n', ' ')


def _to_float(valor) -> Optional[float]:
    if valor is None or str(valor).strip() == '':
        return None
    try:
        return float(str(valor).replace(',', '.').strip())
    except (ValueError, TypeError):
        return None


def _to_str(valor) -> Optional[str]:
    s = str(valor).strip() if valor is not None else ''
    return s if s else None


def _bytes_from_wb(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cabecalho_style(ws, row: int, cols: int):
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='D9D9D9')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = bold
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=False, vertical='center')


# ── Modelos (templates em branco) ─────────────────────────────────────────────

def gerar_modelo_plo() -> bytes:
    """Gera xlsx em branco com o layout esperado para importar a PLO do contrato/obra."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PLO'

    cabecalhos = ['CÓDIGO', 'DESCRIÇÃO', 'UNIDADE', 'QNTD.', 'VALOR TOTAL COM BDI']
    ws.append(cabecalhos)
    _cabecalho_style(ws, 1, len(cabecalhos))

    exemplo_fill = PatternFill('solid', fgColor='F2F2F2')
    ws.append(['3.3.14', 'Exemplo de serviço', 'UN', 10, 15000.00])
    for col in range(1, len(cabecalhos) + 1):
        ws.cell(2, col).fill = exemplo_fill
        ws.cell(2, col).font = Font(italic=True, color='888888')
    ws.cell(2, 4).number_format = '#,##0.00'
    ws.cell(2, 5).number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    return _bytes_from_wb(wb)


def gerar_modelo_cotacoes() -> bytes:
    """Gera xlsx em branco com o layout de importação de cotações."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cotações'

    cabecalhos = [
        'CÓDIGO',
        'FORNECEDOR A', 'VALOR A',
        'FORNECEDOR B', 'VALOR B',
        'FORNECEDOR C', 'VALOR C',
        'OBSERVAÇÃO',
    ]
    ws.append(cabecalhos)
    _cabecalho_style(ws, 1, len(cabecalhos))

    exemplo_fill = PatternFill('solid', fgColor='F2F2F2')
    ws.append(['3.3.14', 'Fornecedor XYZ', 150.00, 'Fornecedor ABC', 145.00, '', '', 'Obs. aqui'])
    for col in range(1, len(cabecalhos) + 1):
        ws.cell(2, col).fill = exemplo_fill
        ws.cell(2, col).font = Font(italic=True, color='888888')

    larguras = [14, 28, 12, 28, 12, 28, 12, 30]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return _bytes_from_wb(wb)


# ── Importações ───────────────────────────────────────────────────────────────

def importar_plo_contrato(dados: bytes, contrato_id: int) -> dict:
    """
    Importa PLO do contrato (xlsx) e popula itens_referencia + contrato_itens.
    Colunas obrigatórias: CÓDIGO, DESCRIÇÃO, UNIDADE.
    Colunas QNTD. e VALOR TOTAL COM BDI são lidas mas ignoradas aqui
    (pertencem à importação de obra via importar_plo_obra).
    Retorna {importados, atualizados, erros}.
    """
    resultado = {'importados': 0, 'atualizados': 0, 'erros': []}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        resultado['erros'].append(f'Erro ao abrir arquivo: {e}')
        return resultado

    if not rows:
        resultado['erros'].append('Arquivo vazio.')
        return resultado

    cabecalho = [_normalizar_cabecalho(c) for c in rows[0]]
    mapa = {h: i for i, h in enumerate(cabecalho)}

    def _idx(*opcoes):
        for op in opcoes:
            if op in mapa:
                return mapa[op]
        return None

    idx_cod  = _idx('CÓDIGO', 'CODIGO', 'ITEM', 'COD', 'CÓD')
    idx_desc = _idx('DESCRIÇÃO', 'DESCRICAO', 'DESCRIÇÃO DO SERVIÇO', 'DESCRICÃO')
    idx_und  = _idx('UNIDADE', 'UNID', 'UNID.', 'UND', 'UN')

    if idx_cod is None or idx_desc is None or idx_und is None:
        resultado['erros'].append(
            'Cabeçalho inválido. Colunas obrigatórias: CÓDIGO, DESCRIÇÃO, UNIDADE. '
            f'Encontrado: {list(mapa.keys())}'
        )
        return resultado

    db = CotacoesDatabase()
    codigos_existentes = db.listar_codigos_contrato(contrato_id)

    itens_ref = []
    itens_ci = []

    for n, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        codigo   = _to_str(row[idx_cod]  if idx_cod  < len(row) else None)
        descricao = _to_str(row[idx_desc] if idx_desc < len(row) else None)
        unidade  = _to_str(row[idx_und]  if idx_und  < len(row) else None)

        if not codigo:
            resultado['erros'].append(f'Linha {n}: CÓDIGO vazio, ignorada.')
            continue
        if not descricao:
            resultado['erros'].append(f'Linha {n}: DESCRIÇÃO vazia para código {codigo}, ignorada.')
            continue
        if not unidade:
            unidade = 'UN'

        itens_ref.append({'codigo': codigo, 'descricao': descricao, 'unidade': unidade, 'tipo': 'servico'})
        itens_ci.append({'item_codigo': codigo, 'curva': None})

    if not itens_ref:
        resultado['erros'].append('Nenhum item válido encontrado no arquivo.')
        return resultado

    try:
        db.upsert_itens_referencia(itens_ref)
        db.upsert_contrato_itens(contrato_id, itens_ci)
        codigos_importados = {i['codigo'] for i in itens_ref}
        resultado['importados'] = len(codigos_importados - codigos_existentes)
        resultado['atualizados'] = len(codigos_importados & codigos_existentes)
    except Exception as e:
        log_error(e, "cotacoes_service", f"importar_plo_contrato contrato {contrato_id}")
        resultado['erros'].append(f'Erro ao salvar no banco: {e}')

    return resultado


def importar_template_cotacoes(dados: bytes, contrato_id: int) -> dict:
    """
    Importa template de cotações preenchido e atualiza cotacoes_contrato.
    Âncora: CÓDIGO. Células vazias = NULL (não sobrescreve valor anterior).
    Retorna {atualizados, ignorados, erros}.
    """
    resultado = {'atualizados': 0, 'ignorados': 0, 'erros': []}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        resultado['erros'].append(f'Erro ao abrir arquivo: {e}')
        return resultado

    if not rows:
        resultado['erros'].append('Arquivo vazio.')
        return resultado

    cabecalho = [_normalizar_cabecalho(c) for c in rows[0]]
    mapa = {h: i for i, h in enumerate(cabecalho)}

    def _idx(*opcoes):
        for op in opcoes:
            if op in mapa:
                return mapa[op]
        return None

    idx_cod    = _idx('CÓDIGO', 'CODIGO', 'ITEM', 'CÓD')
    idx_fa     = _idx('FORNECEDOR A', 'FORN. A', 'FORN A')
    idx_va     = _idx('VALOR A', 'VAL. A', 'VAL A')
    idx_fb     = _idx('FORNECEDOR B', 'FORN. B', 'FORN B')
    idx_vb     = _idx('VALOR B', 'VAL. B', 'VAL B')
    idx_fc     = _idx('FORNECEDOR C', 'FORN. C', 'FORN C')
    idx_vc     = _idx('VALOR C', 'VAL. C', 'VAL C')
    idx_obs    = _idx('OBSERVAÇÃO', 'OBSERVACAO', 'OBS', 'OBS.')

    if idx_cod is None:
        resultado['erros'].append(
            f'Cabeçalho inválido. Coluna CÓDIGO não encontrada. Encontrado: {list(mapa.keys())}'
        )
        return resultado

    db = CotacoesDatabase()
    # Códigos válidos do contrato para validação
    itens_validos = {i['codigo'] for i in db.listar_itens_contrato(contrato_id)}

    cotacoes = []
    for n, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        def _get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        codigo = _to_str(_get(idx_cod))
        if not codigo:
            continue

        if codigo not in itens_validos:
            resultado['ignorados'] += 1
            resultado['erros'].append(f'Linha {n}: código {codigo} não pertence ao contrato, ignorado.')
            continue

        cot = {
            'item_codigo':      codigo,
            'cot_a_fornecedor': _to_str(_get(idx_fa)),
            'cot_a_valor':      _to_float(_get(idx_va)),
            'cot_b_fornecedor': _to_str(_get(idx_fb)),
            'cot_b_valor':      _to_float(_get(idx_vb)),
            'cot_c_fornecedor': _to_str(_get(idx_fc)),
            'cot_c_valor':      _to_float(_get(idx_vc)),
            'observacao':       _to_str(_get(idx_obs)),
        }
        cotacoes.append(cot)

    if not cotacoes:
        resultado['erros'].append('Nenhuma linha válida encontrada.')
        return resultado

    try:
        db.upsert_cotacoes_batch(contrato_id, cotacoes)
        resultado['atualizados'] = len(cotacoes)
    except Exception as e:
        log_error(e, "cotacoes_service", f"importar_template_cotacoes contrato {contrato_id}")
        resultado['erros'].append(f'Erro ao salvar no banco: {e}')

    return resultado


def importar_plo_obra(dados: bytes, obra_id: int) -> dict:
    """
    Importa PLO da obra (xlsx) e popula abc_obra com quantitativos.
    Colunas obrigatórias: CÓDIGO.
    Opcionais: DESCRIÇÃO, UNIDADE, QNTD., VALOR TOTAL COM BDI.
    A CURVA é calculada automaticamente no momento da exportação ABC.
    Retorna {importados, atualizados, erros}.
    """
    resultado = {'importados': 0, 'atualizados': 0, 'erros': []}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        resultado['erros'].append(f'Erro ao abrir arquivo: {e}')
        return resultado

    if not rows:
        resultado['erros'].append('Arquivo vazio.')
        return resultado

    cabecalho = [_normalizar_cabecalho(c) for c in rows[0]]
    mapa = {h: i for i, h in enumerate(cabecalho)}

    def _idx(*opcoes):
        for op in opcoes:
            if op in mapa:
                return mapa[op]
        return None

    idx_cod  = _idx('CÓDIGO', 'CODIGO', 'ITEM', 'CÓD')
    idx_desc = _idx('DESCRIÇÃO', 'DESCRICAO')
    idx_und  = _idx('UNIDADE', 'UNID', 'UNID.', 'UND', 'UN')
    idx_qtd  = _idx('QNTD.', 'QUANTIDADE', 'QTDE', 'QTD', 'QTDE.', 'QNTD')
    idx_bdi  = _idx('VALOR TOTAL COM BDI', 'TOTAL COM BDI', 'VALOR TOTAL', 'TOTAL BDI')

    if idx_cod is None:
        resultado['erros'].append(
            f'Cabeçalho inválido. Coluna CÓDIGO não encontrada. Encontrado: {list(mapa.keys())}'
        )
        return resultado

    itens = []
    for n, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        def _get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        codigo = _to_str(_get(idx_cod))
        if not codigo:
            continue

        itens.append({
            'item_codigo':   codigo,
            'descricao':     _to_str(_get(idx_desc)),
            'unidade':       _to_str(_get(idx_und)) or 'UN',
            'quantidade':    _to_float(_get(idx_qtd)),
            'total_com_bdi': _to_float(_get(idx_bdi)),
            'curva':         None,  # calculada no export ABC
        })

    if not itens:
        resultado['erros'].append('Nenhuma linha válida encontrada.')
        return resultado

    try:
        db = CotacoesDatabase()
        codigos_existentes = db.listar_abc_codigos(obra_id)
        db.upsert_abc_itens(obra_id, itens)
        codigos_importados = {i['item_codigo'] for i in itens}
        resultado['importados'] = len(codigos_importados - codigos_existentes)
        resultado['atualizados'] = len(codigos_importados & codigos_existentes)
    except Exception as e:
        log_error(e, "cotacoes_service", f"importar_plo_obra obra {obra_id}")
        resultado['erros'].append(f'Erro ao salvar no banco: {e}')

    return resultado


# ── Exportações ───────────────────────────────────────────────────────────────

def exportar_template_cotacoes(contrato_id: int) -> bytes:
    """
    Gera xlsx com os itens do contrato + cotações existentes pré-preenchidas.
    Colunas: CÓDIGO | DESCRIÇÃO | UNIDADE | FORNECEDOR A | VALOR A | ... | OBSERVAÇÃO
    """
    db = CotacoesDatabase()
    itens = db.listar_itens_contrato(contrato_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template Cotações'

    cabecalhos = [
        'CÓDIGO', 'DESCRIÇÃO', 'UNIDADE',
        'FORNECEDOR A', 'VALOR A',
        'FORNECEDOR B', 'VALOR B',
        'FORNECEDOR C', 'VALOR C',
        'OBSERVAÇÃO',
    ]
    ws.append(cabecalhos)
    _cabecalho_style(ws, 1, len(cabecalhos))

    for item in itens:
        ws.append([
            item.get('codigo', ''),
            item.get('descricao', ''),
            item.get('unidade', ''),
            item.get('cot_a_fornecedor') or '',
            item.get('cot_a_valor'),
            item.get('cot_b_fornecedor') or '',
            item.get('cot_b_valor'),
            item.get('cot_c_fornecedor') or '',
            item.get('cot_c_valor'),
            item.get('observacao') or '',
        ])

    # Formata colunas de valor como número (cols 5, 7, 9)
    for col in (5, 7, 9):
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = '#,##0.00'

    larguras = [14, 50, 10, 28, 12, 28, 12, 28, 12, 30]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    return _bytes_from_wb(wb)


def _calcular_curva_abc(itens: list) -> list:
    """
    Classifica os itens em curva A/B/C pelo valor total acumulado.
    Thresholds clássicos: A ≤ 70%, B ≤ 90%, C > 90%.
    Retorna a lista com campo 'curva' preenchido em cada item.
    """
    total_geral = sum(i.get('total_com_bdi') or 0 for i in itens)
    if total_geral == 0:
        for i in itens:
            i['curva'] = ''
        return itens

    # Ordena por total_com_bdi decrescente para calcular acumulado
    ordenados = sorted(itens, key=lambda i: i.get('total_com_bdi') or 0, reverse=True)
    acumulado = 0.0
    mapa_curva = {}
    for item in ordenados:
        acumulado += item.get('total_com_bdi') or 0
        pct = acumulado / total_geral
        if pct <= 0.70:
            curva = 'A'
        elif pct <= 0.90:
            curva = 'B'
        else:
            curva = 'C'
        mapa_curva[item['item_codigo']] = curva

    for item in itens:
        item['curva'] = mapa_curva.get(item['item_codigo'], '')
    return itens


def exportar_abc(obra_id: int, contrato_id: int) -> bytes:
    """
    Gera planilha ABC no formato de referência.
    A CURVA (A/B/C) é calculada automaticamente pelo percentual acumulado do
    VALOR TOTAL COM BDI: A ≤ 70%, B ≤ 90%, C > 90%.
    """
    db = CotacoesDatabase()
    abc_itens = db.listar_abc_obra(obra_id)
    cotacoes = {c['item_codigo']: c for c in db.listar_cotacoes_contrato(contrato_id)}

    # Calcula CURVA dinamicamente
    abc_itens = _calcular_curva_abc(abc_itens)

    # Ordena por curva A→B→C e depois por total_com_bdi desc dentro de cada curva
    _ordem_curva = {'A': 0, 'B': 1, 'C': 2, '': 3}
    abc_itens.sort(key=lambda i: (
        _ordem_curva.get(i.get('curva', ''), 3),
        -(i.get('total_com_bdi') or 0)
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ABC'

    cabecalhos = [
        'CURVA', 'ITEM', 'DESCRIÇÃO', 'UNID.', 'QTDE', 'TOTAL COM BDI',
        'COTAÇÃO 01', 'VALOR 01',
        'COTAÇÃO 02', 'VALOR 02',
        'COTAÇÃO 03', 'VALOR 03',
        'MÉDIA', 'OBSERVAÇÃO',
    ]
    ws.append(cabecalhos)
    _cabecalho_style(ws, 1, len(cabecalhos))

    thin = Side(style='thin')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    fmt_num = '#,##0.00'

    # Cores de preenchimento por curva
    fill_curva = {
        'A': PatternFill('solid', fgColor='E8F5E9'),  # verde claro
        'B': PatternFill('solid', fgColor='FFF8E1'),  # amarelo claro
        'C': PatternFill('solid', fgColor='FBE9E7'),  # laranja claro
    }

    for item in abc_itens:
        codigo = item['item_codigo']
        cot = cotacoes.get(codigo, {})
        curva = item.get('curva') or ''

        qtde = item.get('quantidade')
        va = cot.get('cot_a_valor')
        vb = cot.get('cot_b_valor')
        vc = cot.get('cot_c_valor')

        valores = [v for v in (va, vb, vc) if v is not None]
        media = round(sum(valores) / len(valores), 4) if valores else None

        row_data = [
            curva,
            codigo,
            item.get('descricao') or '',
            item.get('unidade') or '',
            qtde,
            item.get('total_com_bdi'),
            cot.get('cot_a_fornecedor') or '',
            va,
            cot.get('cot_b_fornecedor') or '',
            vb,
            cot.get('cot_c_fornecedor') or '',
            vc,
            media,
            cot.get('observacao') or '',
        ]
        ws.append(row_data)
        row_num = ws.max_row

        fill = fill_curva.get(curva)
        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row_num, col)
            cell.border = borda
            if fill:
                cell.fill = fill

        for col in (5, 6, 8, 10, 12, 13):
            ws.cell(row_num, col).number_format = fmt_num

        # Curva em negrito
        ws.cell(row_num, 1).font = Font(bold=True)

    larguras = [8, 14, 50, 8, 10, 14, 28, 12, 28, 12, 28, 12, 12, 30]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    return _bytes_from_wb(wb)
